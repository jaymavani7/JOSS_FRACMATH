# -*- coding: utf-8 -*-
# =====================================================================
# run_3pb_all.py
#
# SINGLE Python file for the notched 3-point bending test.
# Does everything end-to-end:
#
#   1. Builds the geometry (CPS3 free-triangle mesh)
#   2. Writes the .inp deck
#   3. Submits Abaqus with the UMAT (cdm_umat_2d.for)
#   4. Extracts load + CMOD history from the .odb
#   5. Extracts the damage field at the final frame
#   6. Reads wall-clock from the .sta and peak RAM from the .dat
#   7. Plots:  load-vs-CMOD, mesh visualization, damage contour
#   8. Saves EVERYTHING into one Excel file with multiple sheets
#
# The script auto-detects which environment it is in and dispatches
# itself in three phases:
#
#   plain python    -> driver       (orchestrate: cae, job, extract)
#   abaqus cae      -> phase_build  (build + write .inp)
#   abaqus python   -> phase_extract (extract + plot + Excel)
#
# USAGE (single command, from a normal shell):
#     python3 run_3pb_all.py
#
# Output (in ./Gregoire_3PB/results/):
#     load_vs_cmod.png/pdf            curve
#     mesh.png/pdf                    mesh + supports + load
#     damage_contour.png/pdf          omega field at final frame
#     Gregoire_3PB_results.xlsx       one file, multiple sheets:
#         Summary       peak load + CMOD + wall-clock + RAM
#         Load_CMOD     full history of CMOD and Load
#         Damage_Field  element coords + omega
#         Timing        breakdown if available from .sta
#
# Same data as the MATLAB reference:
#     Gregoire D=100, a/D=0.2 (FifthNotched), Medium CPS3 mesh
#     E=37000 MPa, nu=0.20, ft=3.5 MPa, GF=0.090 N/mm, fc/ft=10
#
# Authors: [Name to be added]
# =====================================================================

import os
import sys
import subprocess
import time as _time


# =====================================================================
# CONFIG (shared by all phases)
# =====================================================================
MODEL    = 'Gregoire_3PB'

# Geometry (mm)
D         = 100.0
S         = 2.5 * D
overhang  = 0.5 * D
L         = S + 2.0 * overhang
b_thick   = 50.0
a0        = 0.2 * D
wn        = D / 40.0
xC        = L / 2.0
xL_notch  = xC - wn/2.0
xR_notch  = xC + wn/2.0

# Mesh
ELEM_SIZE_GLOBAL = D / 8.0           # 12.5 mm
ELEM_SIZE_REFINE = ELEM_SIZE_GLOBAL / 4.0   # 3.125 mm
REFINE_W = 0.5 * D
REFINE_H = D

# Material (matches MATLAB default_params)
E_C, NU_C, FT, GF, FCFT = 37000.0, 0.20, 3.5, 0.090, 10.0

# Loading
U_FINAL = -0.2    # mm, downward midspan deflection
N_INC   = 10000

# Output paths
CWD       = os.getcwd()
CASE_DIR  = os.path.join(CWD, MODEL)
RES_DIR   = os.path.join(CASE_DIR, 'results')
UMAT_PATH = os.path.join(CWD, 'cdm_umat_2d.for')


# =====================================================================
# PHASE A: build the model + write the .inp  (runs inside abaqus cae)
# =====================================================================
def phase_build():
    from abaqus import mdb
    from abaqusConstants import (TWO_D_PLANAR, DEFORMABLE_BODY, SIDE1,
                                  FINER, FREE, TRI, ADVANCING_FRONT,
                                  CARTESIAN, ON, OFF,
                                  CPS3, STANDARD, SET)
    import mesh as abqMesh

    if not os.path.exists(CASE_DIR):
        os.makedirs(CASE_DIR)

    if MODEL in mdb.models.keys():
        del mdb.models[MODEL]
    m = mdb.Model(name=MODEL)

    # --- beam sketch with the notch slot at the bottom centre ---------
    sk = m.ConstrainedSketch(name='BeamSketch', sheetSize=2.0*L)
    sk.Line(point1=(0.0,0.0),       point2=(xL_notch,0.0))
    sk.Line(point1=(xL_notch,0.0),  point2=(xL_notch,a0))
    sk.Line(point1=(xL_notch,a0),   point2=(xR_notch,a0))
    sk.Line(point1=(xR_notch,a0),   point2=(xR_notch,0.0))
    sk.Line(point1=(xR_notch,0.0),  point2=(L,0.0))
    sk.Line(point1=(L,0.0), point2=(L,D))
    sk.Line(point1=(L,D),   point2=(0.0,D))
    sk.Line(point1=(0.0,D), point2=(0.0,0.0))

    part = m.Part(name='Beam2D', dimensionality=TWO_D_PLANAR,
                   type=DEFORMABLE_BODY)
    part.BaseShell(sketch=sk)
    del sk

    # --- partition the refinement zone --------------------------------
    tform = part.MakeSketchTransform(sketchPlane=part.faces[0],
                                      sketchPlaneSide=SIDE1,
                                      origin=(0.0,0.0,0.0))
    sk_r = m.ConstrainedSketch(name='RefineZone', sheetSize=L,
                                transform=tform)
    sk_r.rectangle(point1=(xC - REFINE_W/2.0, 0.0),
                   point2=(xC + REFINE_W/2.0, REFINE_H))
    part.PartitionFaceBySketch(sketch=sk_r, faces=part.faces)
    del sk_r

    # --- material + section -------------------------------------------
    mat = m.Material(name='Concrete')
    mat.UserMaterial(mechanicalConstants=(E_C, NU_C, FT, GF, FCFT))
    mat.Depvar(n=2)
    m.HomogeneousSolidSection(name='BeamSec', material='Concrete',
                              thickness=b_thick)
    part.SectionAssignment(region=(part.faces,), sectionName='BeamSec')

    # --- mesh ---------------------------------------------------------
    elem_t3 = abqMesh.ElemType(elemCode=CPS3, elemLibrary=STANDARD)
    part.setElementType(regions=(part.faces,), elemTypes=(elem_t3,))
    for f in part.faces:
        part.setMeshControls(regions=(f,), technique=FREE,
                              elemShape=TRI, algorithm=ADVANCING_FRONT)
    part.seedPart(size=ELEM_SIZE_GLOBAL, deviationFactor=0.1)
    refine_edges = part.edges.getByBoundingBox(
        xMin=xC - REFINE_W/2.0 - 1e-3, xMax=xC + REFINE_W/2.0 + 1e-3,
        yMin=-1e-3, yMax=REFINE_H + 1e-3)
    if len(refine_edges):
        part.seedEdgeBySize(edges=refine_edges, size=ELEM_SIZE_REFINE,
                             constraint=FINER)
    part.generateMesh()
    print('  mesh: %d CPS3 elements, %d nodes'
          % (len(part.elements), len(part.nodes)))

    # --- node sets ----------------------------------------------------
    def pick_single_nearest(p, xt, yt):
        best, bd = None, 1e30
        for nd in p.nodes:
            x,y = nd.coordinates[0], nd.coordinates[1]
            d = (x-xt)**2 + (y-yt)**2
            if d < bd: bd, best = d, nd
        return p.nodes.sequenceFromLabels(labels=(best.label,))

    def pick_n_nearest_top(p, xt, y_top, n):
        tol = 1e-3*abs(y_top) + 1e-6
        cands = []
        for nd in p.nodes:
            x,y = nd.coordinates[0], nd.coordinates[1]
            if abs(y - y_top) > tol: continue
            cands.append((abs(x-xt), nd.label))
        cands.sort()
        keep = [lb for _,lb in cands[:max(1,n)]]
        return p.nodes.sequenceFromLabels(labels=tuple(keep))

    node_tol = 0.5 * ELEM_SIZE_REFINE

    sup_L      = pick_single_nearest(part, overhang,    0.0)
    sup_R      = pick_single_nearest(part, L-overhang,  0.0)
    load_nodes = pick_n_nearest_top(part, xC, D, 3)
    cmod_L     = part.nodes.getByBoundingBox(
        xMin=xL_notch-node_tol, xMax=xL_notch+node_tol,
        yMin=-node_tol,         yMax=node_tol)
    cmod_R     = part.nodes.getByBoundingBox(
        xMin=xR_notch-node_tol, xMax=xR_notch+node_tol,
        yMin=-node_tol,         yMax=node_tol)

    part.Set(nodes=sup_L,      name='Support_Left')
    part.Set(nodes=sup_R,      name='Support_Right')
    part.Set(nodes=load_nodes, name='Load_Nodes')
    part.Set(nodes=cmod_L,     name='CMOD1')
    part.Set(nodes=cmod_R,     name='CMOD2')

    # --- assembly + BCs ------------------------------------------------
    asm = m.rootAssembly
    asm.DatumCsysByDefault(CARTESIAN)
    inst = asm.Instance(name='BeamInst', part=part, dependent=ON)
    for nm in ('Support_Left','Support_Right','Load_Nodes','CMOD1','CMOD2'):
        asm.Set(nodes=inst.sets[nm].nodes, name=nm)

    m.StaticStep(name='Loading', previous='Initial',
                  maxNumInc=4000, initialInc=1.0/N_INC,
                  minInc=1e-10, maxInc=1.0/N_INC, nlgeom=OFF)
    m.DisplacementBC(name='BC_SupL', createStepName='Loading',
                      region=asm.sets['Support_Left'], u1=SET, u2=SET)
    m.DisplacementBC(name='BC_SupR', createStepName='Loading',
                      region=asm.sets['Support_Right'], u2=SET)
    m.DisplacementBC(name='BC_Load', createStepName='Loading',
                      region=asm.sets['Load_Nodes'], u2=U_FINAL)

    m.FieldOutputRequest(name='F-Output', createStepName='Loading',
                          variables=('U','RF','S','E','SDV'), frequency=10)
    m.HistoryOutputRequest(name='H-Load',  createStepName='Loading',
                            variables=('U2','RF2'),
                            region=asm.sets['Load_Nodes'])
    m.HistoryOutputRequest(name='H-CMOD1', createStepName='Loading',
                            variables=('U1',), region=asm.sets['CMOD1'])
    m.HistoryOutputRequest(name='H-CMOD2', createStepName='Loading',
                            variables=('U1',), region=asm.sets['CMOD2'])

    if MODEL in mdb.jobs.keys():
        del mdb.jobs[MODEL]
    job = mdb.Job(name=MODEL, model=MODEL,
                   description='Gregoire D=100 a/D=0.2 3PB, CDM UMAT')

    cwd0 = os.getcwd()
    os.chdir(CASE_DIR)
    try:
        job.writeInput(consistencyChecking=OFF)
    finally:
        os.chdir(cwd0)
    print('  .inp written: ' + os.path.join(CASE_DIR, MODEL + '.inp'))


# =====================================================================
# PHASE B: extract + plot + Excel  (runs inside abaqus python)
# =====================================================================
def phase_extract():
    from odbAccess import openOdb

    odb_path = os.path.join(CASE_DIR, MODEL + '.odb')
    sta_path = os.path.join(CASE_DIR, MODEL + '.sta')
    dat_path = os.path.join(CASE_DIR, MODEL + '.dat')
    if not os.path.exists(odb_path):
        print('ERROR: ODB not found: ' + odb_path)
        return 1

    if not os.path.isdir(RES_DIR):
        os.makedirs(RES_DIR)

    odb = openOdb(odb_path, readOnly=True)
    step = odb.steps['Loading']
    asm  = odb.rootAssembly

    # ---- 1. Load-CMOD history -----------------------------------------
    def gather_labels(nset_name):
        out = set()
        for nlist in asm.nodeSets[nset_name].nodes:
            for nd in nlist: out.add(nd.label)
        return out

    load_lbls  = gather_labels('LOAD_NODES')
    cmod1_lbls = gather_labels('CMOD1')
    cmod2_lbls = gather_labels('CMOD2')

    def reduce_series(labels, key, reducer='sum'):
        series = None; n = 0
        for hkey, hreg in step.historyRegions.items():
            if not hkey.startswith('Node '): continue
            try: lbl = int(hkey.split('.')[-1])
            except: continue
            if lbl not in labels: continue
            if key not in hreg.historyOutputs: continue
            data = hreg.historyOutputs[key].data
            if series is None:
                series = [(t, v) for (t, v) in data]
            else:
                for i, (t, v) in enumerate(data):
                    series[i] = (series[i][0], series[i][1] + v)
            n += 1
        if series is None or n == 0:
            return None
        if reducer == 'avg':
            series = [(t, v/n) for (t, v) in series]
        return series

    rf = reduce_series(load_lbls,  'RF2', 'sum')
    u1 = reduce_series(cmod1_lbls, 'U1',  'avg')
    u2 = reduce_series(cmod2_lbls, 'U1',  'avg')
    if rf is None or u1 is None or u2 is None:
        print('ERROR: missing history outputs in ODB.')
        odb.close()
        return 1

    rows = []
    for i in range(len(rf)):
        t_val = rf[i][0]
        load  = -rf[i][1]                # reaction; positive = downward
        cmod  = abs(u2[i][1] - u1[i][1])
        rows.append((t_val, cmod, load))
    print('  extracted %d history points' % len(rows))

    # peak
    pk_idx, pk = 0, 0.0
    for i, (_, c, ld) in enumerate(rows):
        if ld > pk: pk, pk_idx = ld, i
    pk_cmod = rows[pk_idx][1]

    # ---- 2. Damage field at last frame --------------------------------
    inst = odb.rootAssembly.instances['BEAMINST']

    # Build node coord map for mesh + element centroid coords
    node_map = {}
    for nd in inst.nodes:
        node_map[nd.label] = (nd.coordinates[0], nd.coordinates[1])

    elem_conn = []
    elem_lbls = []
    for el in inst.elements:
        elem_conn.append(el.connectivity)
        elem_lbls.append(el.label)

    last_frame = step.frames[-1]
    print('  damage extracted at frame time = %.4f' % last_frame.frameValue)

    sdv2 = None
    for key in ('SDV2', 'SDV_2'):
        if key in last_frame.fieldOutputs:
            sdv2 = last_frame.fieldOutputs[key]; break

    elem_damage = {}
    if sdv2 is not None:
        for v in sdv2.values:
            d_val = float(v.data) if not hasattr(v.data, '__len__') \
                    else float(sum(v.data)/len(v.data))
            # average across IPs if multiple per element
            if v.elementLabel in elem_damage:
                elem_damage[v.elementLabel] = \
                    0.5*(elem_damage[v.elementLabel] + d_val)
            else:
                elem_damage[v.elementLabel] = d_val

    # Build per-element centroid + damage arrays
    centroids = []
    for i, conn in enumerate(elem_conn):
        xs = [node_map[n][0] for n in conn if n in node_map]
        ys = [node_map[n][1] for n in conn if n in node_map]
        cx, cy = sum(xs)/len(xs), sum(ys)/len(ys)
        d = elem_damage.get(elem_lbls[i], 0.0)
        centroids.append((elem_lbls[i], cx, cy, d))

    # ---- 3. Timing + memory from .sta / .dat --------------------------
    wall = None
    if os.path.exists(sta_path):
        with open(sta_path) as fh:
            lines = fh.readlines()
        for ln in reversed(lines):
            up = ln.upper()
            if 'WALLCLOCK' in up:
                for tok in ln.split():
                    try:
                        wall = float(tok); break
                    except ValueError: pass
                if wall is not None: break

    ram_mb = None
    if os.path.exists(dat_path):
        with open(dat_path) as fh:
            for ln in fh:
                up = ln.upper()
                if 'MEMORY TO MINIMIZE' in up or \
                   ('MEMORY' in up and 'MB' in up) or \
                   'PEAK MEMORY' in up:
                    for tok in ln.replace(',', ' ').split():
                        try:
                            val = float(tok)
                            if 5 < val < 1e6:
                                ram_mb = val
                                break
                        except ValueError: pass
                    if ram_mb is not None: break

    odb.close()

    # ---- 4. Write CSV (always, regardless of Excel) -------------------
    csv_path = os.path.join(RES_DIR, 'load_cmod_data.csv')
    with open(csv_path, 'w') as fh:
        fh.write('time, cmod_mm, load_N\n')
        for (t,c,l) in rows:
            fh.write('%.6e, %.6e, %.6e\n' % (t,c,l))
    print('  wrote ' + csv_path)

    # ---- 5. Plots ------------------------------------------------------
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import numpy as np

        cmods_arr = np.array([r[1] for r in rows])
        loads_arr = np.array([r[2] for r in rows]) / 1000.0   # kN

        # Plot 1: load vs CMOD
        fig, ax = plt.subplots(figsize=(6.5, 4.5))
        ax.plot(cmods_arr, loads_arr, '-',
                color='#1f4e79', lw=2.0, label='Abaqus + UMAT')
        ax.plot(pk_cmod, pk/1000.0,
                marker='*', linestyle='None', markersize=20,
                markerfacecolor='#FFD700', markeredgecolor='black',
                label='Peak: %.2f kN' % (pk/1000.0))
        ax.annotate('  Peak = %.2f kN\n  CMOD = %.4f mm'
                     % (pk/1000.0, pk_cmod),
                     xy=(pk_cmod, pk/1000.0),
                     xytext=(pk_cmod + cmods_arr[-1]*0.15,
                             (pk/1000.0) * 0.80),
                     fontsize=10,
                     arrowprops=dict(arrowstyle='->', color='gray'))
        ax.set_xlabel('CMOD [mm]', fontsize=11)
        ax.set_ylabel('Load [kN]', fontsize=11)
        ax.set_title('3-Point Bending: Load vs CMOD')
        ax.set_xlim(0, max(cmods_arr.max(), 1e-3)*1.05)
        ax.set_ylim(0, max(loads_arr.max(), 1e-3)*1.18)
        ax.grid(True, alpha=0.3, linestyle=':')
        ax.legend(loc='upper right')
        plt.tight_layout()
        plt.savefig(os.path.join(RES_DIR, 'load_vs_cmod.png'), dpi=200)
        plt.savefig(os.path.join(RES_DIR, 'load_vs_cmod.pdf'))
        plt.close(fig)
        print('  wrote load_vs_cmod.png/pdf')

        # Plot 2: mesh + supports + load
        fig, ax = plt.subplots(figsize=(9, 4))
        for conn in elem_conn:
            pts = [node_map[n] for n in conn if n in node_map]
            if len(pts) < 3: continue
            xs = [p[0] for p in pts] + [pts[0][0]]
            ys = [p[1] for p in pts] + [pts[0][1]]
            ax.plot(xs, ys, '-', color='#7a8aa8', lw=0.25)
        # supports
        for ns in ('SUPPORT_LEFT','SUPPORT_RIGHT'):
            for nlist in asm.nodeSets[ns].nodes:
                for nd in nlist:
                    if nd.label in node_map:
                        x, y = node_map[nd.label]
                        ax.plot([x-5, x+5, x, x-5], [y-6, y-6, y, y-6],
                                '-', color='k', lw=1.2)
                        ax.fill([x-5, x+5, x], [y-6, y-6, y],
                                color='#3c3c3c')
        # load
        for nlist in asm.nodeSets['LOAD_NODES'].nodes:
            for nd in nlist:
                if nd.label in node_map:
                    x, y = node_map[nd.label]
                    ax.plot(x, y, 'o', color='#0e7c66', markersize=6)
        load_xs = [node_map[nd.label][0]
                    for nlist in asm.nodeSets['LOAD_NODES'].nodes
                    for nd in nlist if nd.label in node_map]
        if load_xs:
            xC_ = sum(load_xs)/len(load_xs)
            ax.annotate('', xy=(xC_, D+1), xytext=(xC_, D+18),
                        arrowprops=dict(arrowstyle='-|>',
                                        color='#0e7c66', lw=2.2))
            ax.text(xC_+10, D+10, '$P$', fontsize=12,
                    color='#0e7c66', fontweight='bold')
        ax.set_aspect('equal')
        ax.set_xlim(-10, L+10)
        ax.set_ylim(-15, D+25)
        ax.set_xlabel('x [mm]'); ax.set_ylabel('y [mm]')
        ax.set_title('Mesh + supports + load')
        plt.tight_layout()
        plt.savefig(os.path.join(RES_DIR, 'mesh.png'), dpi=200)
        plt.savefig(os.path.join(RES_DIR, 'mesh.pdf'))
        plt.close(fig)
        print('  wrote mesh.png/pdf')

        # Plot 3: damage contour
        cx_arr = np.array([c[1] for c in centroids])
        cy_arr = np.array([c[2] for c in centroids])
        d_arr  = np.array([c[3] for c in centroids])

        fig, ax = plt.subplots(figsize=(9, 4))
        for conn in elem_conn:
            pts = [node_map[n] for n in conn if n in node_map]
            if len(pts) < 3: continue
            xs = [p[0] for p in pts] + [pts[0][0]]
            ys = [p[1] for p in pts] + [pts[0][1]]
            ax.plot(xs, ys, '-', color='#cccccc', lw=0.15)
        sc = ax.scatter(cx_arr, cy_arr, c=d_arr, s=14,
                         cmap='hot_r', vmin=0.0, vmax=1.0,
                         linewidths=0)
        if (d_arr >= 0.95).any():
            mask = d_arr >= 0.95
            ax.scatter(cx_arr[mask], cy_arr[mask], c='red',
                        marker='x', s=18,
                        label='omega >= 0.95 (%d el)' % mask.sum())
            ax.legend(loc='upper right', fontsize=9)
        cb = fig.colorbar(sc, ax=ax, shrink=0.8)
        cb.set_label('damage $\\omega$')
        ax.set_aspect('equal')
        ax.set_xlim(-10, L+10); ax.set_ylim(-5, D+5)
        ax.set_xlabel('x [mm]'); ax.set_ylabel('y [mm]')
        ax.set_title('Damage field at final frame')
        plt.tight_layout()
        plt.savefig(os.path.join(RES_DIR, 'damage_contour.png'), dpi=200)
        plt.savefig(os.path.join(RES_DIR, 'damage_contour.pdf'))
        plt.close(fig)
        print('  wrote damage_contour.png/pdf')

    except ImportError:
        print('  WARNING: matplotlib not available in this Abaqus python '
              '-- plots skipped.')

    # ---- 6. Excel (xlsxwriter -> openpyxl -> CSV fallback) ------------
    xlsx_path = os.path.join(RES_DIR, MODEL + '_results.xlsx')
    written_via = write_excel(xlsx_path, MODEL, rows, centroids,
                                pk, pk_cmod, wall, ram_mb,
                                len(node_map), len(elem_conn))
    print('  wrote %s  (via %s)' % (xlsx_path, written_via))

    # ---- 7. Console summary -------------------------------------------
    print('')
    print('=' * 60)
    print('Summary:')
    print('  Peak load : %.2f N' % pk)
    print('  CMOD@peak : %.4f mm' % pk_cmod)
    if wall   is not None: print('  Wall-clock: %.1f s' % wall)
    if ram_mb is not None: print('  Peak RAM  : %.1f MB' % ram_mb)
    print('  Output    -> ' + RES_DIR)
    print('=' * 60)
    return 0


# =====================================================================
# Excel writer (3 fallbacks)
# =====================================================================
def write_excel(path, model, rows, centroids,
                pk_load, pk_cmod, wall, ram_mb, n_nodes, n_elems):
    """Try xlsxwriter, then openpyxl, then CSV (multi-file) fallback.
    Returns the name of the backend actually used."""

    # ---- Try xlsxwriter -----------------------------------------------
    try:
        import xlsxwriter
        wb = xlsxwriter.Workbook(path)

        bold   = wb.add_format({'bold': True})
        title  = wb.add_format({'bold': True, 'font_size': 14,
                                'bg_color': '#1F4E79', 'font_color': 'white'})
        hdr    = wb.add_format({'bold': True, 'bg_color': '#DCE6F1',
                                'border': 1, 'align': 'center'})
        cell   = wb.add_format({'border': 1})
        num    = wb.add_format({'border': 1, 'num_format': '0.000000'})

        # Sheet 1: Summary
        ws = wb.add_worksheet('Summary')
        ws.merge_range('A1:B1', model + '  --  Abaqus 3PB Results', title)
        ws.set_column(0, 0, 22); ws.set_column(1, 1, 18)
        rows_summary = [
            ('Model',           model),
            ('Mesh element',    'CPS3 (3-node plane stress)'),
            ('Number of nodes', n_nodes),
            ('Number of elem',  n_elems),
            ('E [MPa]',         E_C),
            ('nu',              NU_C),
            ('ft [MPa]',        FT),
            ('GF [N/mm]',       GF),
            ('fc/ft',           FCFT),
            ('Span S [mm]',     S),
            ('Depth D [mm]',    D),
            ('Notch a0 [mm]',   a0),
            ('Notch wn [mm]',   wn),
            ('Peak load [N]',   pk_load),
            ('CMOD@peak [mm]',  pk_cmod),
            ('Wall-clock [s]',  wall if wall is not None else 'N/A'),
            ('Peak RAM [MB]',   ram_mb if ram_mb is not None else 'N/A'),
        ]
        ws.write(2, 0, 'Quantity', hdr); ws.write(2, 1, 'Value', hdr)
        for i, (k, v) in enumerate(rows_summary):
            ws.write(3 + i, 0, k, cell)
            ws.write(3 + i, 1, v, cell)

        # Sheet 2: Load_CMOD
        ws2 = wb.add_worksheet('Load_CMOD')
        ws2.set_column(0, 2, 14)
        ws2.write(0, 0, 'time',    hdr)
        ws2.write(0, 1, 'cmod_mm', hdr)
        ws2.write(0, 2, 'load_N',  hdr)
        for i, (t, c, ld) in enumerate(rows):
            ws2.write_number(i+1, 0, t,  num)
            ws2.write_number(i+1, 1, c,  num)
            ws2.write_number(i+1, 2, ld, num)

        # Add chart
        ch = wb.add_chart({'type': 'line'})
        ch.add_series({
            'name':       'Load-CMOD',
            'categories': ['Load_CMOD', 1, 1, len(rows), 1],
            'values':     ['Load_CMOD', 1, 2, len(rows), 2],
            'line':       {'color': '#1F4E79', 'width': 2.0},
        })
        ch.set_title({'name': 'Load vs CMOD'})
        ch.set_x_axis({'name': 'CMOD [mm]'})
        ch.set_y_axis({'name': 'Load [N]'})
        ch.set_legend({'none': True})
        ws2.insert_chart('E2', ch, {'x_scale': 1.2, 'y_scale': 1.2})

        # Sheet 3: Damage_Field
        ws3 = wb.add_worksheet('Damage_Field')
        ws3.set_column(0, 3, 14)
        ws3.write(0, 0, 'elem_label', hdr)
        ws3.write(0, 1, 'x_mm',       hdr)
        ws3.write(0, 2, 'y_mm',       hdr)
        ws3.write(0, 3, 'omega',      hdr)
        for i, (lbl, cx, cy, d) in enumerate(centroids):
            ws3.write_number(i+1, 0, lbl, cell)
            ws3.write_number(i+1, 1, cx,  num)
            ws3.write_number(i+1, 2, cy,  num)
            ws3.write_number(i+1, 3, d,   num)

        # Sheet 4: Timing
        ws4 = wb.add_worksheet('Timing')
        ws4.set_column(0, 1, 22)
        ws4.write(0, 0, 'Metric', hdr); ws4.write(0, 1, 'Value', hdr)
        ws4.write(1, 0, 'Wall-clock [s]', cell)
        ws4.write(1, 1, wall if wall is not None else 'N/A', cell)
        ws4.write(2, 0, 'Peak RAM [MB]', cell)
        ws4.write(2, 1, ram_mb if ram_mb is not None else 'N/A', cell)
        ws4.write(3, 0, 'Mesh DOFs',  cell)
        ws4.write(3, 1, 2*n_nodes,    cell)
        ws4.write(4, 0, 'Mesh elems', cell)
        ws4.write(4, 1, n_elems,      cell)

        wb.close()
        return 'xlsxwriter'

    except ImportError:
        pass

    # ---- Try openpyxl --------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()

        thin = Side(style='thin', color='888888')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill('solid', fgColor='DCE6F1')
        hdr_font = Font(bold=True)
        title_font = Font(bold=True, size=14, color='FFFFFF')
        title_fill = PatternFill('solid', fgColor='1F4E79')

        ws = wb.active; ws.title = 'Summary'
        ws.merge_cells('A1:B1')
        c = ws['A1']
        c.value = model + ' -- Abaqus 3PB Results'
        c.font = title_font; c.fill = title_fill
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws['A3'] = 'Quantity'; ws['B3'] = 'Value'
        for col in ('A3','B3'):
            ws[col].font = hdr_font; ws[col].fill = hdr_fill
        rows_summary = [
            ('Model', model), ('Mesh element', 'CPS3'),
            ('Number of nodes', n_nodes), ('Number of elem', n_elems),
            ('E [MPa]', E_C), ('nu', NU_C),
            ('ft [MPa]', FT), ('GF [N/mm]', GF), ('fc/ft', FCFT),
            ('Span S [mm]', S), ('Depth D [mm]', D),
            ('Notch a0 [mm]', a0), ('Notch wn [mm]', wn),
            ('Peak load [N]', pk_load), ('CMOD@peak [mm]', pk_cmod),
            ('Wall-clock [s]', wall if wall is not None else 'N/A'),
            ('Peak RAM [MB]',  ram_mb if ram_mb is not None else 'N/A'),
        ]
        for i,(k,v) in enumerate(rows_summary):
            ws.cell(row=4+i, column=1, value=k).border = border
            ws.cell(row=4+i, column=2, value=v).border = border

        ws2 = wb.create_sheet('Load_CMOD')
        for j, h in enumerate(['time','cmod_mm','load_N'], start=1):
            cc = ws2.cell(row=1, column=j, value=h)
            cc.font = hdr_font; cc.fill = hdr_fill
        for i, (t, c, ld) in enumerate(rows):
            ws2.cell(row=i+2, column=1, value=t)
            ws2.cell(row=i+2, column=2, value=c)
            ws2.cell(row=i+2, column=3, value=ld)

        ws3 = wb.create_sheet('Damage_Field')
        for j, h in enumerate(['elem_label','x_mm','y_mm','omega'], start=1):
            cc = ws3.cell(row=1, column=j, value=h)
            cc.font = hdr_font; cc.fill = hdr_fill
        for i, (lbl, cx, cy, d) in enumerate(centroids):
            ws3.cell(row=i+2, column=1, value=lbl)
            ws3.cell(row=i+2, column=2, value=cx)
            ws3.cell(row=i+2, column=3, value=cy)
            ws3.cell(row=i+2, column=4, value=d)

        ws4 = wb.create_sheet('Timing')
        for j, h in enumerate(['Metric','Value'], start=1):
            cc = ws4.cell(row=1, column=j, value=h)
            cc.font = hdr_font; cc.fill = hdr_fill
        for i, (k, v) in enumerate([
            ('Wall-clock [s]', wall if wall is not None else 'N/A'),
            ('Peak RAM [MB]',  ram_mb if ram_mb is not None else 'N/A'),
            ('Mesh DOFs',  2*n_nodes),
            ('Mesh elems', n_elems),
        ]):
            ws4.cell(row=i+2, column=1, value=k)
            ws4.cell(row=i+2, column=2, value=v)

        wb.save(path)
        return 'openpyxl'

    except ImportError:
        pass

    # ---- Final fallback: multi-sheet via 4 CSVs -----------------------
    base = path.replace('.xlsx', '')
    with open(base + '_Summary.csv', 'w') as fh:
        fh.write('Quantity,Value\n')
        for k, v in [
            ('Model', model), ('Mesh element', 'CPS3'),
            ('Number of nodes', n_nodes), ('Number of elem', n_elems),
            ('E [MPa]', E_C), ('nu', NU_C),
            ('ft [MPa]', FT), ('GF [N/mm]', GF), ('fc/ft', FCFT),
            ('Peak load [N]', pk_load), ('CMOD@peak [mm]', pk_cmod),
            ('Wall-clock [s]', wall if wall is not None else 'N/A'),
            ('Peak RAM [MB]',  ram_mb if ram_mb is not None else 'N/A'),
        ]:
            fh.write('%s,%s\n' % (k, v))
    with open(base + '_Load_CMOD.csv', 'w') as fh:
        fh.write('time,cmod_mm,load_N\n')
        for t, c, ld in rows:
            fh.write('%g,%g,%g\n' % (t, c, ld))
    with open(base + '_Damage_Field.csv', 'w') as fh:
        fh.write('elem_label,x_mm,y_mm,omega\n')
        for lbl, cx, cy, d in centroids:
            fh.write('%d,%g,%g,%g\n' % (lbl, cx, cy, d))
    return 'csv_fallback'


# =====================================================================
# DRIVER (plain python orchestrator)
# =====================================================================
def driver():
    print('=' * 60)
    print('  Abaqus 3PB driver -- single Python file')
    print('  Geometry: D=100 mm, a/D=0.2, span=250 mm')
    print('  Material: E=37 GPa, ft=3.5 MPa, GF=0.090 N/mm, fc/ft=10')
    print('=' * 60)

    if not os.path.exists(UMAT_PATH):
        sys.exit('ERROR: cdm_umat_2d.for not found in ' + CWD)

    # --- Phase A: build (call self under abaqus cae) -------------------
    print('\n[1/3] Building model + writing .inp...')
    rc = subprocess.call(['abaqus', 'cae',
                          'noGUI=' + os.path.abspath(__file__),
                          '--', 'build'])
    if rc != 0:
        sys.exit('Abaqus CAE build failed.')

    # --- Run the job ---------------------------------------------------
    print('\n[2/3] Submitting Abaqus job (this is the slow part)...')
    cwd0 = os.getcwd(); os.chdir(CASE_DIR)
    t0 = _time.time()
    try:
        rc = subprocess.call(['abaqus', 'job=' + MODEL,
                              'user=' + UMAT_PATH,
                              'interactive', 'cpus=4',
                              'ask_delete=OFF'])
    finally:
        os.chdir(cwd0)
    elapsed = _time.time() - t0
    if rc != 0:
        print('WARNING: Abaqus job exited with status %d. Trying to '
              'extract whatever is in the .odb anyway.' % rc)
    else:
        print('  Abaqus job finished in %.1f s' % elapsed)

    # --- Phase B: extract + plot + Excel -------------------------------
    print('\n[3/3] Extracting + plotting + writing Excel...')
    rc = subprocess.call(['abaqus', 'python',
                          os.path.abspath(__file__),
                          '--', 'extract'])
    if rc != 0:
        sys.exit('Extraction phase failed.')

    print('\nDone. See ' + RES_DIR + '/ for outputs.')


# =====================================================================
# Mode dispatch
# =====================================================================
def _mode():
    argv = ' '.join(sys.argv)
    if ' build' in argv or sys.argv[-1] == 'build':
        return 'build'
    if ' extract' in argv or sys.argv[-1] == 'extract':
        return 'extract'
    try:
        import abaqus           # noqa: F401
        return 'build'
    except ImportError:
        return 'driver'


if __name__ == '__main__':
    mode = _mode()
    if   mode == 'driver':  driver()
    elif mode == 'build':   phase_build()
    elif mode == 'extract': sys.exit(phase_extract())